import csv
import json
from django.http import JsonResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import transaction
from .models import DataEntry
from django.db.models import Q
import io
import time
import threading
from django.core.paginator import Paginator
from openpyxl import load_workbook

# Global variable to store upload progress
upload_progress = {
    'total_rows': 0,
    'processed_rows': 0,
    'status': 'idle',  # 'idle', 'in_progress', 'completed', 'error'
    'error': None,
    'columns': []
}


def reset_progress():
    """Reset the upload progress tracker"""
    global upload_progress
    upload_progress = {
        'total_rows': 0,
        'processed_rows': 0,
        'status': 'idle',
        'error': None,
        'columns': []
    }


def process_csv_in_background(file_content, delimiter):
    """Process CSV file in the background"""
    global upload_progress

    try:
        # Reset existing progress
        reset_progress()

        # Mark as in progress
        upload_progress['status'] = 'in_progress'

        # Count total rows first (needed for progress tracking)
        total_rows = 0
        for _ in csv.reader(io.StringIO(file_content), delimiter=delimiter):
            total_rows += 1

        # Subtract 1 for header row
        upload_progress['total_rows'] = total_rows - 1 if total_rows > 0 else 0

        # Reset file pointer
        csv_file = io.StringIO(file_content)
        csv_reader = csv.DictReader(csv_file, delimiter=delimiter)

        # Get columns
        columns = csv_reader.fieldnames if csv_reader.fieldnames else []
        upload_progress['columns'] = columns

        # Clear existing data
        with transaction.atomic():
            DataEntry.objects.all().delete()

        # Process in batches
        batch_size = 1000
        batch = []

        for row in csv_reader:
            # Clean up row data
            clean_row = {k: str(v).strip() if v is not None else '' for k, v in row.items()}
            batch.append(DataEntry(data=clean_row))

            # When batch is full, insert and update progress
            if len(batch) >= batch_size:
                with transaction.atomic():
                    DataEntry.objects.bulk_create(batch)

                upload_progress['processed_rows'] += len(batch)
                batch = []

        # Insert any remaining records
        if batch:
            with transaction.atomic():
                DataEntry.objects.bulk_create(batch)
            upload_progress['processed_rows'] += len(batch)

        # Mark as completed
        upload_progress['status'] = 'completed'

    except Exception as e:
        # Handle any errors
        upload_progress['status'] = 'error'
        upload_progress['error'] = str(e)


@csrf_exempt
@require_http_methods(["POST"])
def upload_file(request):
    try:
        file = request.FILES['file']
        file_name = file.name.lower()
        
        # Reset progress tracker
        global upload_progress
        upload_progress = {
            'total_rows': 0,
            'processed_rows': 0,
            'status': 'in_progress',
            'error': None,
            'columns': []
        }
        
        # Clear existing data
        DataEntry.objects.all().delete()
        
        # Initialize variables
        chunk_size = 1000  # Process 1000 rows at a time
        rows_to_create = []
        columns = []
        total_rows_processed = 0
        
        if file_name.endswith('.csv'):
            # Handle CSV file in chunks
            content = b''
            for chunk in file.chunks(8192):
                content += chunk
            
            # Decode the content
            decoded_file = content.decode('utf-8')
            
            # Count total rows first
            total_rows = sum(1 for line in io.StringIO(decoded_file)) - 1  # Subtract header row
            upload_progress['total_rows'] = total_rows
            
            # Process the file
            csv_reader = csv.DictReader(io.StringIO(decoded_file))
            columns = [col.strip() for col in csv_reader.fieldnames if col and col.strip()]
            upload_progress['columns'] = columns
            
            # Process rows in chunks
            for row in csv_reader:
                cleaned_row = {
                    k.strip(): v.strip() if v and isinstance(v, str) else v 
                    for k, v in row.items() 
                    if k and k.strip()
                }
                
                if cleaned_row:
                    rows_to_create.append(DataEntry(data=cleaned_row))
                    total_rows_processed += 1
                    upload_progress['processed_rows'] = total_rows_processed
                
                # When we reach chunk_size, bulk create and reset
                if len(rows_to_create) >= chunk_size:
                    DataEntry.objects.bulk_create(rows_to_create)
                    rows_to_create = []
                    
        elif file_name.endswith('.xlsx'):
            # Handle XLSX file
            wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            first_row = next(ws.rows)
            for cell in first_row:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            columns = [col for col in headers if col]
            upload_progress['columns'] = columns
            
            # Count total rows (excluding header)
            upload_progress['total_rows'] = ws.max_row - 1
            
            # Process data rows in chunks
            for row in ws.rows:
                # Skip header row
                if row[0].row == 1:
                    continue
                    
                row_data = {}
                for header, cell in zip(headers, row):
                    if header and cell.value is not None:
                        value = cell.value
                        # Convert to string if it's a number or date
                        if not isinstance(value, str):
                            value = str(value)
                        row_data[header] = value.strip()
                
                if row_data:
                    rows_to_create.append(DataEntry(data=row_data))
                    total_rows_processed += 1
                    upload_progress['processed_rows'] = total_rows_processed
                
                # When we reach chunk_size, bulk create and reset
                if len(rows_to_create) >= chunk_size:
                    DataEntry.objects.bulk_create(rows_to_create)
                    rows_to_create = []
            
            # Close the workbook to free memory
            wb.close()
        else:
            return JsonResponse({'error': 'Please upload a CSV or XLSX file'}, status=400)
        
        # Bulk create any remaining entries
        if rows_to_create:
            DataEntry.objects.bulk_create(rows_to_create)
        
        # Update progress status
        upload_progress['status'] = 'completed'
        upload_progress['total_rows_processed'] = total_rows_processed
        
        return JsonResponse({
            'message': 'File uploaded successfully',
            'columns': columns,
            'total_rows_processed': total_rows_processed
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Upload error: {error_details}")  # Log the full error
        
        # Update progress status
        upload_progress['status'] = 'error'
        upload_progress['error'] = str(e)
        
        return JsonResponse({
            'error': str(e),
            'details': error_details
        }, status=400)


@csrf_exempt
@require_http_methods(["GET"])
def get_upload_progress(request):
    """Get the current progress of an active upload"""
    global upload_progress
    return JsonResponse(upload_progress)


@csrf_exempt
@require_http_methods(["GET"])
def get_columns(request):
    """Get available columns from the database"""
    try:
        # First check if we have in-progress upload with columns
        global upload_progress
        if upload_progress['status'] in ['in_progress', 'completed'] and upload_progress['columns']:
            return JsonResponse({'columns': upload_progress['columns']})

        # If not, get from database
        sample_entry = DataEntry.objects.first()
        if not sample_entry:
            return JsonResponse({'columns': []})

        columns = list(sample_entry.data.keys())
        return JsonResponse({'columns': columns})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["GET"])
def search_suggestions(request):
    """Get search suggestions based on key columns"""
    try:
        prefix = request.GET.get('q', '').strip()
        if not prefix or len(prefix) < 2:
            return JsonResponse({'suggestions': []})

        # Priority fields to search for suggestions
        priority_fields = ['Product', 'IndianCompany', 'ForeignCompany', 'Indian Company', 'Foreign Company']

        # Build query to find matches in priority fields
        query = Q()
        for field in priority_fields:
            query |= Q(**{f"data__{field}__istartswith": prefix})

        # Get distinct values that match
        suggestions = []
        limit = 10  # Limit number of suggestions

        # Process each priority field
        for field in priority_fields:
            # Skip if we already have enough suggestions
            if len(suggestions) >= limit:
                break

            # Get entries matching the prefix in this field
            matches = DataEntry.objects.filter(**{f"data__{field}__istartswith": prefix}).distinct()

            # Extract unique values for this field
            field_values = set()
            for entry in matches:
                if field in entry.data and entry.data[field]:
                    value = str(entry.data[field]).strip()
                    if value.lower().startswith(prefix.lower()):
                        field_values.add(value)

            # Add values to suggestions (with field info)
            for value in list(field_values)[:limit - len(suggestions)]:
                if value not in [s['value'] for s in suggestions]:
                    suggestions.append({
                        'value': value,
                        'field': field,
                        'display': f"{value} ({field})"
                    })

            # Stop if we have enough suggestions
            if len(suggestions) >= limit:
                break

        return JsonResponse({'suggestions': suggestions})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def search_data(request):
    """Search data with pagination and relevance sorting"""
    try:
        data = json.loads(request.body)
        fields = data.get('fields', [])
        value = data.get('value', '')
        page = data.get('page', 1)
        page_size = data.get('page_size', 20)  # Default to 20 results per page

        if not fields:
            return JsonResponse({'error': 'Fields are required'}, status=400)

        # Handle empty search value
        if not value:
            return JsonResponse({
                'results': [],
                'total_count': 0,
                'page': page,
                'total_pages': 0
            })

        # Identify priority fields for boosting relevance
        priority_fields = []
        normal_fields = []

        # Priority fields mapping (case-insensitive)
        priority_field_names = ['product', 'indiancompany', 'foreigncompany']

        for field in fields:
            clean_field = field.strip()
            normalized = clean_field.lower().replace(' ', '')

            if normalized in priority_field_names:
                priority_fields.append(clean_field)
            else:
                normal_fields.append(clean_field)

        # Build query with two parts: priority and normal
        # This allows us to annotate and order results by relevance
        priority_query = Q()
        for field in priority_fields:
            priority_query |= Q(**{f"data__{field}__icontains": value})

        normal_query = Q()
        for field in normal_fields:
            normal_query |= Q(**{f"data__{field}__icontains": value})

        # Combine queries
        combined_query = priority_query | normal_query

        # Perform search
        results = DataEntry.objects.filter(combined_query)

        # Apply relevance ranking: prioritize matches in important fields
        # We'll use a custom sorting approach by adding annotations
        # For PostgreSQL, we could use more advanced features

        # First get all matching results
        all_results = list(results)

        # Define a function to calculate relevance score
        def calculate_relevance(entry_data):
            score = 0
            search_value = value.lower()

            # Check priority fields first (higher score)
            for field in priority_fields:
                if field in entry_data and entry_data[field]:
                    field_value = str(entry_data[field]).lower()
                    
                    # Exact match gets highest score
                    if field_value == search_value:
                        if field.lower() == 'product':
                            score += 200  # Product exact match
                        else:
                            score += 150  # Company exact match
                    # Prefix match gets high score
                    elif field_value.startswith(search_value):
                        if field.lower() == 'product':
                            score += 150  # Product prefix match
                        else:
                            score += 100  # Company prefix match
                    # Contains match gets base priority score
                    elif search_value in field_value:
                        if field.lower() == 'product':
                            score += 100  # Product contains match
                        else:
                            score += 50   # Company contains match

            # Then check other fields (lower priority)
            for field in normal_fields:
                if field in entry_data and entry_data[field]:
                    field_value = str(entry_data[field]).lower()
                    if field_value == search_value:
                        score += 30  # Exact match in other fields
                    elif field_value.startswith(search_value):
                        score += 20  # Prefix match in other fields
                    elif search_value in field_value:
                        score += 10  # Contains match in other fields

            return score

        # Sort results by relevance score (higher first)
        sorted_results = sorted(
            all_results,
            key=lambda entry: calculate_relevance(entry.data),
            reverse=True
        )

        # Get total count
        total_count = len(sorted_results)

        # Manual pagination
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_results = sorted_results[start_idx:end_idx]
        total_pages = (total_count + page_size - 1) // page_size  # Ceiling division

        # Prepare response with additional relevance info
        response_data = {
            'results': [
                {
                    **entry.data,
                    '_relevance': calculate_relevance(entry.data)  # Add relevance score for debugging
                }
                for entry in paginated_results
            ],
            'total_count': total_count,
            'page': page,
            'total_pages': total_pages
        }

        return JsonResponse(response_data, safe=False)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["GET"])
def download_results(request):
    """Download search results as CSV"""
    try:
        fields = request.GET.getlist('fields')
        value = request.GET.get('value', '')

        if not fields:
            return JsonResponse({'error': 'Fields are required'}, status=400)

        # Build query
        query = Q()
        for field in fields:
            clean_field = field.strip()
            query |= Q(**{f"data__{clean_field}__icontains": value})

        # Perform search
        results = DataEntry.objects.filter(query)

        if not results.exists():
            return JsonResponse({'error': 'No results to download'}, status=404)

        # Get all column names from the first result
        first_entry = results.first()
        fieldnames = list(first_entry.data.keys())

        # Create CSV writer
        def generate_csv():
            buffer = io.StringIO()
            writer = csv.DictWriter(buffer, fieldnames=fieldnames)

            # Write header
            writer.writeheader()
            yield buffer.getvalue()
            buffer.seek(0)
            buffer.truncate(0)

            # Write rows in batches
            batch_size = 100
            for i in range(0, results.count(), batch_size):
                batch = results[i:i + batch_size]
                for entry in batch:
                    writer.writerow(entry.data)
                    yield buffer.getvalue()
                    buffer.seek(0)
                    buffer.truncate(0)

        # Create streaming response
        response = StreamingHttpResponse(
            generate_csv(),
            content_type='text/csv'
        )
        response['Content-Disposition'] = 'attachment; filename="search_results.csv"'

        return response

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)