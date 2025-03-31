import csv
import io
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import transaction
from ..models import DataEntry, FileInfo
from openpyxl import load_workbook
from .state import upload_progress, active_file_id, reset_progress

@csrf_exempt
@require_http_methods(["POST"])
def upload_file(request):
    try:
        file = request.FILES['file']
        file_name = file.name.lower()
        
        # Reset progress tracker
        global upload_progress
        upload_progress = {
            'total_rows': 0,
            'processed_rows': 0,
            'status': 'in_progress',
            'error': None,
            'columns': [],
            'current_file': file.name
        }
        
        # Create new FileInfo entry
        file_info = FileInfo.objects.create(
            filename=file.name,
            is_active=True
        )
        
        # Initialize variables
        chunk_size = 1000  # Process 1000 rows at a time
        rows_to_create = []
        columns = []
        total_rows_processed = 0
        
        if file_name.endswith('.csv'):
            # Handle CSV file in chunks
            content = b''
            for chunk in file.chunks(8192):
                content += chunk
            
            # Decode the content
            decoded_file = content.decode('utf-8')
            
            # Count total rows first
            total_rows = sum(1 for line in io.StringIO(decoded_file)) - 1  # Subtract header row
            upload_progress['total_rows'] = total_rows
            
            # Process the file
            csv_reader = csv.DictReader(io.StringIO(decoded_file))
            columns = [col.strip() for col in csv_reader.fieldnames if col and col.strip()]
            upload_progress['columns'] = columns
            
            # Process rows in chunks
            for row in csv_reader:
                cleaned_row = {
                    k.strip(): v.strip() if v and isinstance(v, str) else v 
                    for k, v in row.items() 
                    if k and k.strip()
                }
                
                if cleaned_row:
                    rows_to_create.append(DataEntry(data=cleaned_row, file=file_info))
                    total_rows_processed += 1
                    upload_progress['processed_rows'] = total_rows_processed
                
                # When we reach chunk_size, bulk create and reset
                if len(rows_to_create) >= chunk_size:
                    DataEntry.objects.bulk_create(rows_to_create)
                    rows_to_create = []
                    
        elif file_name.endswith('.xlsx'):
            # Handle XLSX file
            wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            first_row = next(ws.rows)
            for cell in first_row:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            columns = [col for col in headers if col]
            upload_progress['columns'] = columns
            
            # Count total rows (excluding header)
            upload_progress['total_rows'] = ws.max_row - 1
            
            # Process data rows in chunks
            for row in ws.rows:
                # Skip header row
                if row[0].row == 1:
                    continue
                    
                row_data = {}
                for header, cell in zip(headers, row):
                    if header and cell.value is not None:
                        value = cell.value
                        # Convert to string if it's a number or date
                        if not isinstance(value, str):
                            value = str(value)
                        row_data[header] = value.strip()
                
                if row_data:
                    rows_to_create.append(DataEntry(data=row_data, file=file_info))
                    total_rows_processed += 1
                    upload_progress['processed_rows'] = total_rows_processed
                
                # When we reach chunk_size, bulk create and reset
                if len(rows_to_create) >= chunk_size:
                    DataEntry.objects.bulk_create(rows_to_create)
                    rows_to_create = []
            
            # Close the workbook to free memory
            wb.close()
        else:
            return JsonResponse({'error': 'Please upload a CSV or XLSX file'}, status=400)
        
        # Bulk create any remaining entries
        if rows_to_create:
            DataEntry.objects.bulk_create(rows_to_create)
        
        # Update file info with final row count
        file_info.row_count = total_rows_processed
        file_info.save()
        
        # Set this as the active file
        global active_file_id
        active_file_id = file_info.id
        
        # Update progress status
        upload_progress['status'] = 'completed'
        upload_progress['total_rows_processed'] = total_rows_processed
        
        return JsonResponse({
            'message': 'File uploaded successfully',
            'file_id': file_info.id,
            'columns': columns,
            'total_rows_processed': total_rows_processed
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Upload error: {error_details}")
        
        upload_progress['status'] = 'error'
        upload_progress['error'] = str(e)
        
        return JsonResponse({
            'error': str(e),
            'details': error_details
        }, status=400) 